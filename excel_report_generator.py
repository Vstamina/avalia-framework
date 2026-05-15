# excel_report_generator.py

from io import BytesIO
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter

from framework_engine import (
    AssessmentResult,
    summarize_assessment,
    get_maturity_ruler_rows,
)


# ==========================================================
# ESTILOS
# ==========================================================

PRIMARY = "0FA3B1"
PRIMARY_DARK = "082B63"
ACCENT = "59C3C3"
BG_LIGHT = "F5F7FA"
CARD_BG = "FFFFFF"
TEXT_DARK = "1F2937"
TEXT_MUTED = "6B7280"
BORDER = "D9E2EC"
DANGER = "E76F51"
WARNING = "F4A261"
SUCCESS = "2E8B57"


def _thin_border():
    side = Side(style="thin", color=BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _header_style(cell, fill=PRIMARY_DARK, font_color="FFFFFF"):
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(bold=True, color=font_color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _thin_border()


def _body_style(cell, wrap=True):
    cell.font = Font(color=TEXT_DARK)
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)
    cell.border = _thin_border()


def _title_style(cell):
    cell.font = Font(size=18, bold=True, color=PRIMARY_DARK)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _subtitle_style(cell):
    cell.font = Font(size=11, color=TEXT_MUTED)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _format_number(value):
    try:
        return round(float(value), 2)
    except Exception:
        return value


def _safe_text(value):
    if value is None:
        return ""
    return str(value)


def _auto_width(ws, max_width=55):
    for col in ws.columns:
        column = col[0].column
        letter = get_column_letter(column)
        max_len = 0

        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))

        ws.column_dimensions[letter].width = min(max_len + 3, max_width)


def _freeze_and_filter(ws, header_row=1):
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = ws.dimensions


# ==========================================================
# ABA 1 — DASHBOARD
# ==========================================================

def _add_dashboard(ws, client_data, results: List[AssessmentResult], summary):
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    ws["A1"] = "Dashboard Executivo — Framework AVALIA"
    _title_style(ws["A1"])

    ws.merge_cells("A2:H2")
    ws["A2"] = (
        "Painel de síntese do diagnóstico de maturidade, prioridades de intervenção "
        "e distribuição das dimensões avaliadas."
    )
    _subtitle_style(ws["A2"])

    ws["A4"] = "Organização"
    ws["B4"] = client_data.get("client_name", "")
    ws["D4"] = "Unidade"
    ws["E4"] = client_data.get("unit", "")
    ws["G4"] = "Período"
    ws["H4"] = client_data.get("period", "")

    for cell in ["A4", "D4", "G4"]:
        _header_style(ws[cell], fill=ACCENT, font_color=PRIMARY_DARK)

    for cell in ["B4", "E4", "H4"]:
        _body_style(ws[cell])

    prioridades_altas = len([r for r in results if r.prioridade_intervencao == "Alta"])
    dimensoes_avaliadas = len(results)

    kpis = [
        ("Índice Geral", _format_number(summary.indice_geral), summary.nivel_geral),
        ("Camada 1", _format_number(summary.indice_camada_1), summary.nivel_camada_1),
        ("Camada 2", _format_number(summary.indice_camada_2), summary.nivel_camada_2),
        ("Prioridades Altas", prioridades_altas, "Foco imediato"),
        ("Dimensões Avaliadas", dimensoes_avaliadas, "Total analisado"),
    ]

    start_col = 1
    for idx, (label, value, help_text) in enumerate(kpis):
        col = start_col + idx * 2
        top = ws.cell(row=6, column=col)
        bottom = ws.cell(row=7, column=col)

        ws.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col + 1)
        ws.merge_cells(start_row=7, start_column=col, end_row=7, end_column=col + 1)

        top.value = label
        top.fill = PatternFill("solid", fgColor=PRIMARY_DARK)
        top.font = Font(bold=True, color="FFFFFF")
        top.alignment = Alignment(horizontal="center")

        bottom.value = f"{value} | {help_text}"
        bottom.fill = PatternFill("solid", fgColor="FFFFFF")
        bottom.font = Font(bold=True, color=PRIMARY)
        bottom.alignment = Alignment(horizontal="center")

        for row in [6, 7]:
            for c in range(col, col + 2):
                ws.cell(row=row, column=c).border = _thin_border()

    # Tabela auxiliar para gráficos
    ws["A10"] = "Dimensão"
    ws["B10"] = "Nota"
    ws["C10"] = "Prioridade"

    for cell in ws[10]:
        _header_style(cell)

    for row_idx, item in enumerate(results, start=11):
        ws.cell(row=row_idx, column=1).value = item.dimensao
        ws.cell(row=row_idx, column=2).value = _format_number(item.score_final)
        ws.cell(row=row_idx, column=3).value = item.prioridade_intervencao

        for col in range(1, 4):
            _body_style(ws.cell(row=row_idx, column=col))

    # Gráfico de barras
    chart = BarChart()
    chart.title = "Maturidade por Dimensão"
    chart.y_axis.title = "Nota"
    chart.x_axis.title = "Dimensão"
    chart.height = 8
    chart.width = 18

    data = Reference(ws, min_col=2, min_row=10, max_row=10 + len(results))
    cats = Reference(ws, min_col=1, min_row=11, max_row=10 + len(results))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    ws.add_chart(chart, "E10")

    # Distribuição de prioridades
    priority_counts = {}
    for item in results:
        priority_counts[item.prioridade_intervencao] = priority_counts.get(item.prioridade_intervencao, 0) + 1

    ws["A26"] = "Prioridade"
    ws["B26"] = "Quantidade"
    _header_style(ws["A26"])
    _header_style(ws["B26"])

    for idx, (priority, count) in enumerate(priority_counts.items(), start=27):
        ws.cell(row=idx, column=1).value = priority
        ws.cell(row=idx, column=2).value = count
        _body_style(ws.cell(row=idx, column=1))
        _body_style(ws.cell(row=idx, column=2))

    pie = PieChart()
    pie.title = "Distribuição das Prioridades"
    pie.height = 7
    pie.width = 9

    labels = Reference(ws, min_col=1, min_row=27, max_row=26 + len(priority_counts))
    data = Reference(ws, min_col=2, min_row=26, max_row=26 + len(priority_counts))
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)

    ws.add_chart(pie, "E26")

    _auto_width(ws)
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 32


# ==========================================================
# ABA 2 — SÍNTESE DAS DIMENSÕES
# ==========================================================

def _add_dimension_summary(ws, results: List[AssessmentResult]):
    headers = [
        "Camada",
        "Dimensão",
        "Nota",
        "Maturidade",
        "Impacto estratégico",
        "Prioridade",
        "Principal lacuna",
        "Recomendação estratégica",
    ]

    ws.append(headers)

    for cell in ws[1]:
        _header_style(cell)

    for item in results:
        principal_lacuna = item.lacunas[0].lacuna if item.lacunas else ""

        ws.append([
            f"Camada {item.camada}",
            item.dimensao,
            _format_number(item.score_final),
            item.nivel_maturidade,
            item.impacto_estrategico_aplicado,
            item.prioridade_intervencao,
            principal_lacuna,
            item.recomendacao_estrategica,
        ])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            _body_style(cell)

        prioridade = row[5].value
        if prioridade == "Alta":
            row[5].fill = PatternFill("solid", fgColor="FCE4DC")
        elif prioridade == "Média":
            row[5].fill = PatternFill("solid", fgColor="FFF1D6")
        elif prioridade == "Monitoramento":
            row[5].fill = PatternFill("solid", fgColor="D9F3F6")
        elif prioridade == "Manutenção":
            row[5].fill = PatternFill("solid", fgColor="DFF3E7")

    _freeze_and_filter(ws)
    _auto_width(ws)


# ==========================================================
# ABA 3 — LACUNAS E RECOMENDAÇÕES
# ==========================================================

def _add_gaps_recommendations(ws, results: List[AssessmentResult]):
    headers = [
        "Camada",
        "Dimensão",
        "Lacuna identificada",
        "Evidência associada",
        "Impacto gerencial",
        "Prioridade",
        "Recomendação estratégica",
        "Observações do consultor",
    ]

    ws.append(headers)

    for cell in ws[1]:
        _header_style(cell)

    for item in results:
        if item.lacunas:
            for lacuna in item.lacunas:
                ws.append([
                    f"Camada {item.camada}",
                    item.dimensao,
                    lacuna.lacuna,
                    lacuna.evidencia,
                    lacuna.impacto,
                    item.prioridade_intervencao,
                    item.recomendacao_estrategica,
                    "",
                ])
        else:
            ws.append([
                f"Camada {item.camada}",
                item.dimensao,
                "",
                "",
                "",
                item.prioridade_intervencao,
                item.recomendacao_estrategica,
                "",
            ])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            _body_style(cell)

    _freeze_and_filter(ws)
    _auto_width(ws)


# ==========================================================
# ABA 4 — ROTEIRO PARA RELATÓRIO
# ==========================================================

def _add_report_script(ws, client_data, results: List[AssessmentResult], summary):
    headers = [
        "Seção do relatório",
        "Conteúdo sugerido pelo sistema",
        "Observações / texto do consultor",
    ]

    ws.append(headers)

    for cell in ws[1]:
        _header_style(cell)

    prioridades = [
        item.dimensao for item in results
        if item.prioridade_intervencao == "Alta"
    ]

    forcas = summary.dimensoes_maior_maturidade

    rows = [
        [
            "Sumário Executivo",
            (
                f"O diagnóstico AVALIA identificou índice geral de maturidade de "
                f"{_format_number(summary.indice_geral)}, situando a organização no patamar "
                f"{summary.nivel_geral}. A Camada 1 apresentou índice "
                f"{_format_number(summary.indice_camada_1)}, enquanto a Camada 2 apresentou "
                f"{_format_number(summary.indice_camada_2)}."
            ),
            "",
        ],
        [
            "Forças identificadas",
            "; ".join(forcas) if forcas else "Não foram identificadas forças relativas.",
            "",
        ],
        [
            "Prioridades de intervenção",
            "; ".join(prioridades) if prioridades else "Não foram identificadas prioridades altas.",
            "",
        ],
        [
            "Intervenção proposta",
            (
                "Recomenda-se estruturar um modelo integrado de pós-venda e feedback qualificado, "
                "com governança, jornada, indicadores, tratamento de feedbacks e aprendizagem organizacional."
            ),
            "",
        ],
        [
            "Roadmap",
            (
                "Fase 1: alinhamento conceitual; Fase 2: desenho de fluxos e instrumentos; "
                "Fase 3: piloto controlado; Fase 4: avaliação e ajustes; "
                "Fase 5: escala e institucionalização."
            ),
            "",
        ],
        [
            "Conclusão",
            (
                "O diagnóstico indica a necessidade de consolidar capacidades de acompanhamento, "
                "demonstração de valor, fechamento de loop e uso gerencial das evidências."
            ),
            "",
        ],
    ]

    for row in rows:
        ws.append(row)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            _body_style(cell)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 55

    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 80

    _freeze_and_filter(ws)


# ==========================================================
# ABA 5 — BASE TÉCNICA
# ==========================================================

def _add_technical_base(ws):
    ws.append(["Item", "Descrição"])

    for cell in ws[1]:
        _header_style(cell)

    rows = [
        ["Framework", "AVALIA — Diagnóstico de Pós-Venda e Feedback Qualificado"],
        ["Escala", "0 a 3"],
        ["Regra geral", "Quanto menor a maturidade e maior o impacto estratégico, maior a prioridade de intervenção."],
        ["Impacto estratégico", "Pode ser definido automaticamente por dimensão ou ajustado manualmente pelo consultor."],
        ["Prioridade", "Alta, Média, Monitoramento ou Manutenção."],
    ]

    for row in rows:
        ws.append(row)

    ws.append([])
    ws.append(["Régua de maturidade", "", ""])

    for row in get_maturity_ruler_rows():
        ws.append(list(row))

    for row in ws.iter_rows():
        for cell in row:
            _body_style(cell)

    for cell in ws[1]:
        _header_style(cell)

    _auto_width(ws)


# ==========================================================
# FUNÇÃO PRINCIPAL
# ==========================================================

def generate_excel_report(client_data: dict, results: List[AssessmentResult]) -> BytesIO:
    summary = summarize_assessment(results)

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    ws_dashboard = wb.create_sheet("Dashboard Executivo")
    ws_summary = wb.create_sheet("Síntese das Dimensões")
    ws_gaps = wb.create_sheet("Lacunas e Recomendações")
    ws_script = wb.create_sheet("Roteiro para Relatório")
    ws_base = wb.create_sheet("Base Técnica")

    _add_dashboard(ws_dashboard, client_data, results, summary)
    _add_dimension_summary(ws_summary, results)
    _add_gaps_recommendations(ws_gaps, results)
    _add_report_script(ws_script, client_data, results, summary)
    _add_technical_base(ws_base)

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return file_stream